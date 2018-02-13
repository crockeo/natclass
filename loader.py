#!/usr/bin/env python3

import openpyxl
import json

def translate_xlsx():
    wb = openpyxl.load_workbook('FeatureChart.xlsx')
    ws = wb['FEATURE CHART']

    # Getting the features
    feature_map = {}
    for i in range(3, 29):
        feature_map[i] = ws[2][i].value.lower()

    # Getting the feature values for each letter.
    features = {}
    for i in range(3, 107):
        if ws[i][2].value == None or (20 <= i and i <= 25):
            continue

        features[ws[i][2].value] = {}
        for j in range(3, 29):
            features[ws[i][2].value][feature_map[j]] = ws[i][j].value

    output = open('features.json', 'w')
    json.dump(features, output)

if __name__ == '__main__':
    translate_xlsx()
